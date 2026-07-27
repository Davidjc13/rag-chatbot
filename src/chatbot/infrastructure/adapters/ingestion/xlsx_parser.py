"""Parser XLSX: cada hoja es una tabla con cabecera."""

from __future__ import annotations

import io
import logging
from pathlib import Path

from openpyxl import load_workbook

from chatbot.application.services.table_markdown import table_to_markdown
from chatbot.domain.documents import ContentBlock, ContentKind, DocumentFormat, ParsedDocument
from chatbot.domain.exceptions import DocumentParseError
from chatbot.domain.ports import DocumentParserPort

logger = logging.getLogger(__name__)


class XlsxParserAdapter(DocumentParserPort):
    def supports(self, filename: str) -> bool:
        return Path(filename).suffix.lower() in {".xlsx", ".xlsm"}

    def parse(self, *, filename: str, data: bytes) -> ParsedDocument:
        blocks: list[ContentBlock] = []
        try:
            workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
            for sheet in workbook.worksheets:
                rows_raw: list[list[object | None]] = []
                for row in sheet.iter_rows(values_only=True):
                    values = list(row)
                    if any(cell is not None and str(cell).strip() for cell in values):
                        rows_raw.append(values)
                if not rows_raw:
                    continue

                headers = ["" if c is None else str(c) for c in rows_raw[0]]
                data_rows = rows_raw[1:]
                column_names = ", ".join(h for h in headers if h.strip())
                blocks.append(
                    ContentBlock(
                        kind=ContentKind.TEXT,
                        text=(
                            f"La hoja '{sheet.title}' contiene una tabla con "
                            f"{len(data_rows)} filas de datos y columnas: {column_names}"
                        ),
                        metadata={"sheet": sheet.title},
                    )
                )
                markdown = table_to_markdown(
                    headers,
                    data_rows,
                    title=f"Hoja: {sheet.title}",
                )
                blocks.append(
                    ContentBlock(
                        kind=ContentKind.TABLE,
                        text=markdown,
                        metadata={"sheet": sheet.title},
                    )
                )
            workbook.close()
        except Exception as exc:  # noqa: BLE001
            logger.exception("Error parseando XLSX %s", filename)
            raise DocumentParseError(
                f"No se pudo parsear el XLSX: {exc}",
                filename=filename,
            ) from exc

        if not blocks:
            raise DocumentParseError("El XLSX no contiene hojas con datos", filename=filename)

        return ParsedDocument(filename=filename, format=DocumentFormat.XLSX, blocks=blocks)
